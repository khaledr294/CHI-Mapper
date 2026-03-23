"""
CHI Drug Formulary - Automatic Update Tool
Probes CHI website for new editions, downloads xlsx, extracts CSVs,
rebuilds the database, and updates the application automatically.
Includes email notification on successful update.
"""

import os
import re
import json
import tempfile
import shutil
import smtplib
import logging
from email.mime.text import MIMEText
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_FILE = os.path.join(BASE_DIR, 'update_state.json')
TEMPLATE_FILE = os.path.join(BASE_DIR, 'templates', 'index.html')

# URL pattern: edition number + date (DDMonYYYY)
URL_TEMPLATE = (
    "https://www.chi.gov.sa/Style%20Library/IDF_Branding/files/"
    "CHI%20Drug%20Formulary%20Compilation%20and%20Formatting%20-%20%20"
    "ed{edition}_{date}.xlsx"
)

MONTH_ABBRS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Column names expected in the target CSVs (must match data_processor.py)
IND_TARGET_COLS = [
    'INDICATION', 'ICD 10 CODE', 'DRUG PHARMACOLOGICAL CLASS ',
    'DRUG PHARMACOLOGICAL SUBCLASS',
    'DESCRIPTION CODE \n(ACTIVE INGREDIENT- STRENGTH-DOSAGE FORM)',
    'SCIENTIFIC NAME ', 'SCIENTIFIC DESCRIPTION CODE ROOT', 'ATC CODE',
    'PHARMACEUTICAL FORM ', 'PHARMACEUTICAL FORM CODE ROOT',
    'ADMINISTRATION ROUTE', 'STRENGTH ', 'STRENGTH UNIT ',
    'SUBSTITUABLE', 'PRESCRIBING EDITS', 'MDD ADULTS', 'MDD PEDIATRICS',
    'NOTES', 'APPENDIX', 'PATIENT TYPE', 'SFDA REGISTRATION STATUS'
]

SFDA_TARGET_COLS = [
    'RegisterNumber', 'ReferenceNumber', 'Old register Number',
    'Product type', 'DrugType', 'Sub-Type', 'Scientific Name',
    'ScientificDescriptionCodeRoot', 'Trade Name', 'Strength',
    'StrengthUnit', 'PharmaceuticalForm', 'PharmaceuticalFormCodeRoot',
    'AdministrationRoute', 'AtcCode1', 'AtcCode2', 'Size', 'SizeUnit',
    'PackageTypes', 'PackageSize', 'Legal Status', 'Product Control',
    'Distribute area', 'Public price', 'shelfLife', 'Storage conditions',
    'Storage Condition Arabic', 'Marketing Company', 'Marketing Country',
    'Manufacture Name', 'Manufacture Country',
    'Secondry package  manufacture', 'Main Agent', 'Secosnd Agent',
    'Third agent', 'Description Code', 'Authorization Status',
    'Last Update', 'GTIN'
]

# Known column name variations across editions (xlsx name → target CSV name)
SFDA_COL_RENAMES = {
    'DescriptionCode': 'Description Code',
    'Third Agent': 'Third agent',
    '2nd Manufacture Name': 'Secondry package  manufacture',
    '2nd Manufacture Country': None,  # Drop — no target equivalent
    'Marketing Status': None,  # Drop — no target equivalent
}

logger = logging.getLogger('chi_updater')


# ═══════════════════════════════════════════════════════════
# State Management
# ═══════════════════════════════════════════════════════════

def load_state():
    """Load the current update state from JSON file."""
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        'current_edition': 57,
        'date_string': '02Mar2026',
        'file_url': None,
        'last_check': None,
        'last_update': None,
        'update_history': []
    }


def save_state(state):
    """Save the update state to JSON file."""
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=4, ensure_ascii=False)


# ═══════════════════════════════════════════════════════════
# URL Probing
# ═══════════════════════════════════════════════════════════

def _check_url(url):
    """Send a HEAD request to check if a URL exists. Returns url if 200, else None."""
    try:
        resp = requests.head(url, timeout=8, allow_redirects=True)
        if resp.status_code == 200:
            return url
    except requests.RequestException:
        pass
    return None


def _generate_candidate_urls(edition, months_back=4):
    """Generate candidate URLs for a given edition number across recent months."""
    now = datetime.now()
    candidates = []

    for month_offset in range(months_back + 2):  # current month + future + past
        # Try months from (now - months_back) to (now + 1)
        offset = month_offset - months_back
        month = now.month + offset
        year = now.year
        while month < 1:
            month += 12
            year -= 1
        while month > 12:
            month -= 12
            year += 1

        mon_abbr = MONTH_ABBRS[month - 1]
        for day in range(1, 32):
            date_str = f"{day:02d}{mon_abbr}{year}"
            url = URL_TEMPLATE.format(edition=edition, date=date_str)
            candidates.append((url, edition, date_str))

    return candidates


def probe_for_new_edition(current_edition, max_ahead=5):
    """
    Probe CHI website for editions newer than current_edition.
    Returns (url, edition_num, date_str) if found, else None.
    """
    for edition in range(current_edition + 1, current_edition + max_ahead + 1):
        candidates = _generate_candidate_urls(edition)
        logger.info(f"Probing for edition {edition} ({len(candidates)} URLs)...")

        # Use thread pool for parallel HEAD requests (5 threads for gentle probing)
        with ThreadPoolExecutor(max_workers=5) as pool:
            future_to_info = {}
            for url, ed, date_str in candidates:
                future = pool.submit(_check_url, url)
                future_to_info[future] = (url, ed, date_str)

            for future in as_completed(future_to_info):
                result = future.result()
                if result:
                    url, ed, date_str = future_to_info[future]
                    logger.info(f"Found new edition: ed{ed} ({date_str})")
                    # Cancel remaining futures
                    pool.shutdown(wait=False, cancel_futures=True)
                    return url, ed, date_str

        logger.info(f"Edition {edition} not found.")

    return None


# ═══════════════════════════════════════════════════════════
# Download
# ═══════════════════════════════════════════════════════════

def download_edition(url, dest_dir=None):
    """
    Download an xlsx file from url.
    Returns path to downloaded file, or None on failure.
    """
    dest_dir = dest_dir or BASE_DIR
    logger.info(f"Downloading: {url}")

    try:
        resp = requests.get(url, timeout=120, stream=True)
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"Download failed: {e}")
        return None

    # Validate content length (must be > 100KB for a valid formulary)
    content_length = resp.headers.get('Content-Length')
    if content_length and int(content_length) < 100_000:
        logger.error(f"File too small ({content_length} bytes), likely not a valid formulary.")
        return None

    # Write to temp file first, then move
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=dest_dir)
    try:
        with os.fdopen(tmp_fd, 'wb') as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)

        file_size = os.path.getsize(tmp_path)
        if file_size < 100_000:
            logger.error(f"Downloaded file too small ({file_size} bytes).")
            os.unlink(tmp_path)
            return None

        # Validate it's a real xlsx by trying to open with openpyxl
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            wb.close()
        except Exception as e:
            logger.error(f"Downloaded file is not a valid xlsx: {e}")
            os.unlink(tmp_path)
            return None

        # Rename to final name
        edition_match = re.search(r'ed(\d+)_(\w+)\.xlsx', url)
        if edition_match:
            filename = f"CHI Drug Formulary Compilation and Formatting -  ed{edition_match.group(1)}_{edition_match.group(2)}.xlsx"
        else:
            filename = f"CHI_Formulary_{datetime.now().strftime('%Y%m%d')}.xlsx"

        final_path = os.path.join(dest_dir, filename)
        shutil.move(tmp_path, final_path)
        logger.info(f"Downloaded to: {final_path} ({file_size / 1024 / 1024:.1f} MB)")
        return final_path

    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


# ═══════════════════════════════════════════════════════════
# CSV Extraction from xlsx
# ═══════════════════════════════════════════════════════════

def _detect_header_row(ws, marker_text, max_rows=15):
    """Find the row index containing marker_text in column A or B."""
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and marker_text.lower() in str(cell.value).lower():
                return row_idx
    return None


def extract_csvs_from_xlsx(xlsx_path, edition_num, date_str, dest_dir=None):
    """
    Extract Indication and SFDA CSVs from an xlsx file.
    Returns (indication_csv_path, sfda_csv_path) or raises on error.
    """
    dest_dir = dest_dir or BASE_DIR
    logger.info(f"Extracting CSVs from: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if len(sheet_names) < 2:
        wb.close()
        raise ValueError(f"Expected at least 2 sheets, found {len(sheet_names)}: {sheet_names}")

    indication_sheet = sheet_names[0]
    sfda_sheet = sheet_names[1]
    wb.close()

    # ─── Extract Indication Sheet ────────────────────────────
    logger.info(f"Processing Indication sheet: '{indication_sheet}'")

    # Detect header row (look for "INDICATION" text)
    wb_check = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws_ind = wb_check[indication_sheet]
    ind_header_row = _detect_header_row(ws_ind, 'INDICATION')
    wb_check.close()

    if ind_header_row is None:
        raise ValueError("Could not find 'INDICATION' header in first sheet")

    # header param is 0-indexed in pandas, row numbers are 1-indexed
    ind_df = pd.read_excel(xlsx_path, sheet_name=indication_sheet,
                           header=ind_header_row - 1, dtype=str)

    # Drop unnamed/empty columns (padding)
    ind_df = ind_df.loc[:, ~ind_df.columns.str.startswith('Unnamed')]
    ind_df = ind_df.dropna(how='all')

    # Verify column count matches
    if len(ind_df.columns) != len(IND_TARGET_COLS):
        logger.warning(
            f"Indication columns: got {len(ind_df.columns)}, expected {len(IND_TARGET_COLS)}. "
            f"Columns: {list(ind_df.columns)}"
        )
        # Try to map by position if count matches after cleanup
        if len(ind_df.columns) > len(IND_TARGET_COLS):
            ind_df = ind_df.iloc[:, :len(IND_TARGET_COLS)]

    # Rename columns to match target
    ind_df.columns = IND_TARGET_COLS[:len(ind_df.columns)]

    ind_csv_name = f"Indication -  ed{edition_num}_{date_str}.csv"
    ind_csv_path = os.path.join(dest_dir, ind_csv_name)
    ind_df.to_csv(ind_csv_path, index=False)
    logger.info(f"  Indication CSV: {len(ind_df)} rows → {ind_csv_name}")

    # ─── Extract SFDA Sheet ──────────────────────────────────
    logger.info(f"Processing SFDA sheet: '{sfda_sheet}'")

    wb_check2 = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws_sfda = wb_check2[sfda_sheet]
    sfda_header_row = _detect_header_row(ws_sfda, 'RegisterNumber')
    if sfda_header_row is None:
        sfda_header_row = _detect_header_row(ws_sfda, 'Register')
    wb_check2.close()

    if sfda_header_row is None:
        raise ValueError("Could not find 'RegisterNumber' header in SFDA sheet")

    sfda_df = pd.read_excel(xlsx_path, sheet_name=sfda_sheet,
                            header=sfda_header_row - 1, dtype=str)

    # Drop unnamed/empty columns
    sfda_df = sfda_df.loc[:, ~sfda_df.columns.str.startswith('Unnamed')]
    sfda_df = sfda_df.dropna(how='all')

    # Apply known column renames
    rename_map = {}
    for old_name, new_name in SFDA_COL_RENAMES.items():
        if old_name in sfda_df.columns:
            if new_name is None:
                sfda_df = sfda_df.drop(columns=[old_name])
            else:
                rename_map[old_name] = new_name
    if rename_map:
        sfda_df = sfda_df.rename(columns=rename_map)

    # Ensure all target columns exist — add missing ones as empty
    for col in SFDA_TARGET_COLS:
        if col not in sfda_df.columns:
            sfda_df[col] = ''
            logger.info(f"  Added missing column: '{col}'")

    # Reorder to match target and keep only target columns
    sfda_df = sfda_df[SFDA_TARGET_COLS]

    sfda_csv_name = f"SFDA Mapping -  ed{edition_num}_{date_str}.csv"
    sfda_csv_path = os.path.join(dest_dir, sfda_csv_name)
    sfda_df.to_csv(sfda_csv_path, index=False)
    logger.info(f"  SFDA CSV: {len(sfda_df)} rows → {sfda_csv_name}")

    return ind_csv_path, sfda_csv_path


# ═══════════════════════════════════════════════════════════
# Footer Update
# ═══════════════════════════════════════════════════════════

def _format_edition_footer(edition_num, date_str):
    """Convert edition 57 + '02Mar2026' → 'Ed57 - Mar 2026'."""
    m = re.match(r'\d{2}(\w{3})(\d{4})', date_str)
    if m:
        return f"Ed{edition_num} - {m.group(1)} {m.group(2)}"
    return f"Ed{edition_num}"


def update_footer(edition_num, date_str):
    """Update the edition text in the HTML footer."""
    if not os.path.exists(TEMPLATE_FILE):
        logger.warning(f"Template not found: {TEMPLATE_FILE}")
        return False

    with open(TEMPLATE_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    new_text = _format_edition_footer(edition_num, date_str)
    updated, count = re.subn(r'Ed\d+\s*-\s*\w+\s+\d{4}', new_text, content)

    if count == 0:
        logger.warning("Could not find edition pattern in footer to update.")
        return False

    with open(TEMPLATE_FILE, 'w', encoding='utf-8') as f:
        f.write(updated)

    logger.info(f"Footer updated to: {new_text}")
    return True


# ═══════════════════════════════════════════════════════════
# Data Processor Integration
# ═══════════════════════════════════════════════════════════

def update_data_processor_paths(indication_csv, sfda_csv):
    """Update the file path constants in data_processor.py."""
    dp_file = os.path.join(BASE_DIR, 'data_processor.py')
    with open(dp_file, 'r', encoding='utf-8') as f:
        content = f.read()

    ind_name = os.path.basename(indication_csv)
    sfda_name = os.path.basename(sfda_csv)

    content = re.sub(
        r"INDICATION_FILE\s*=\s*os\.path\.join\(BASE_DIR,\s*'[^']+'\)",
        f"INDICATION_FILE = os.path.join(BASE_DIR, '{ind_name}')",
        content
    )
    content = re.sub(
        r"SFDA_FILE\s*=\s*os\.path\.join\(BASE_DIR,\s*'[^']+'\)",
        f"SFDA_FILE = os.path.join(BASE_DIR, '{sfda_name}')",
        content
    )

    with open(dp_file, 'w', encoding='utf-8') as f:
        f.write(content)

    logger.info(f"Updated data_processor.py paths → {ind_name}, {sfda_name}")


# ═══════════════════════════════════════════════════════════
# Email Notification
# ═══════════════════════════════════════════════════════════

def send_update_notification(edition_num, date_str, stats=None):
    """
    Send email notification about a successful update.
    Uses SMTP settings from environment variables.
    """
    recipient = os.environ.get('UPDATE_NOTIFY_EMAIL', '')
    smtp_user = os.environ.get('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASSWORD', '')
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', '587'))

    if not recipient or not smtp_user or not smtp_pass:
        logger.info("Email notification skipped (SMTP not configured).")
        return False

    footer_text = _format_edition_footer(edition_num, date_str)
    stats_text = ""
    if stats:
        stats_text = f"""
إحصائيات القاعدة الجديدة:
  - الأدوية: {stats.get('drugs', 'N/A')}
  - الاستطبابات: {stats.get('indications', 'N/A')}
  - المنتجات (SFDA): {stats.get('products', 'N/A')}
  - الربط دواء-استطباب: {stats.get('mappings', 'N/A')}
  - أكواد ICD-10: {stats.get('icd_codes', 'N/A')}
"""

    body = f"""تم تحديث CHI Drug Formulary بنجاح!

النسخة الجديدة: {footer_text}
التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{stats_text}
---
CHI-Mapper Auto-Updater
"""

    msg = MIMEText(body, 'plain', 'utf-8')
    msg['Subject'] = f'CHI Formulary Updated — {footer_text}'
    msg['From'] = smtp_user
    msg['To'] = recipient

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        logger.info(f"Notification email sent to {recipient}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False


# ═══════════════════════════════════════════════════════════
# Database Stats (post-build verification)
# ═══════════════════════════════════════════════════════════

def get_db_stats():
    """Read summary stats from the built database."""
    import sqlite3
    db_path = os.path.join(BASE_DIR, 'chi_mapper.db')
    if not os.path.exists(db_path):
        return None

    conn = sqlite3.connect(db_path)
    try:
        stats = {}
        for table in ['drugs', 'indications', 'drug_indications', 'products']:
            cur = conn.execute(f'SELECT COUNT(*) FROM {table}')
            stats[table] = cur.fetchone()[0]

        cur = conn.execute('SELECT COUNT(DISTINCT icd_code) FROM indication_icd_codes')
        stats['icd_codes'] = cur.fetchone()[0]

        return {
            'drugs': stats['drugs'],
            'indications': stats['indications'],
            'products': stats['products'],
            'mappings': stats['drug_indications'],
            'icd_codes': stats['icd_codes'],
        }
    except Exception:
        return None
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════
# Main Orchestrator
# ═══════════════════════════════════════════════════════════

def run_update(force_edition=None):
    """
    Full update pipeline:
    1. Check for new edition (or use force_edition)
    2. Download xlsx
    3. Extract CSVs
    4. Rebuild database
    5. Update footer
    6. Save state & notify

    Args:
        force_edition: Tuple (url, edition_num, date_str) to skip probing.

    Returns:
        dict with keys: status ('updated', 'no_update', 'error'), details, edition
    """
    state = load_state()
    now = datetime.now(timezone.utc).isoformat()

    # Step 1: Probe for new edition
    if force_edition:
        url, edition_num, date_str = force_edition
        logger.info(f"Forced update to ed{edition_num}")
    else:
        logger.info(f"Checking for updates (current: ed{state['current_edition']})...")
        result = probe_for_new_edition(state['current_edition'])
        state['last_check'] = now
        save_state(state)

        if result is None:
            logger.info("No new edition found.")
            return {'status': 'no_update', 'details': f"Current ed{state['current_edition']} is latest."}

        url, edition_num, date_str = result

    # Step 2: Download
    xlsx_path = download_edition(url)
    if not xlsx_path:
        return {'status': 'error', 'details': f"Failed to download ed{edition_num}"}

    try:
        # Step 3: Extract CSVs
        ind_csv, sfda_csv = extract_csvs_from_xlsx(xlsx_path, edition_num, date_str)

        # Step 4: Update data_processor.py paths and rebuild database
        update_data_processor_paths(ind_csv, sfda_csv)

        from data_processor import build_database
        build_database(indication_file=ind_csv, sfda_file=sfda_csv)

        # Step 5: Verify database
        stats = get_db_stats()
        if stats and stats['drugs'] < 1000:
            logger.error(f"Database verification failed: only {stats['drugs']} drugs (expected >3000)")
            return {'status': 'error', 'details': f"DB verification failed: {stats['drugs']} drugs"}

        # Step 6: Update footer
        update_footer(edition_num, date_str)

        # Step 7: Save state
        state['current_edition'] = edition_num
        state['date_string'] = date_str
        state['file_url'] = url
        state['last_update'] = now
        state['last_check'] = now
        state['update_history'].append({
            'edition': edition_num,
            'date_string': date_str,
            'updated_at': now,
            'stats': stats,
        })
        save_state(state)

        # Step 8: Notify
        send_update_notification(edition_num, date_str, stats)

        details = f"Updated to ed{edition_num} ({date_str})"
        if stats:
            details += f" — {stats['drugs']} drugs, {stats['indications']} indications, {stats['products']} products"
        logger.info(details)

        return {'status': 'updated', 'details': details, 'edition': edition_num, 'stats': stats}

    except Exception as e:
        logger.error(f"Update failed: {e}")
        return {'status': 'error', 'details': str(e)}


def check_and_update():
    """
    Startup-safe wrapper: checks for updates, falls back silently on any error.
    Returns True if an update was applied, False otherwise.
    """
    try:
        result = run_update()
        return result.get('status') == 'updated'
    except Exception as e:
        logger.error(f"Auto-update check failed (falling back to existing data): {e}")
        return False


# ═══════════════════════════════════════════════════════════
# CLI Entry Point
# ═══════════════════════════════════════════════════════════

if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(name)s] %(levelname)s: %(message)s'
    )

    import sys

    if len(sys.argv) > 1 and sys.argv[1] == '--force':
        # Force re-process current edition (useful for testing extraction)
        state = load_state()
        url = state.get('file_url', '')
        ed = state['current_edition']
        ds = state['date_string']
        print(f"Force re-processing current edition: ed{ed} ({ds})")
        result = run_update(force_edition=(url, ed, ds))
    else:
        result = run_update()

    print(f"\nResult: {json.dumps(result, indent=2, ensure_ascii=False)}")
